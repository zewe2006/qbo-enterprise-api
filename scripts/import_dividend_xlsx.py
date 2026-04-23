#!/usr/bin/env python3
"""
One-shot importer for the Sweet Hut dividend spreadsheet.

Parses:
  - `Data` sheet → creates shareholders (if missing) and posts dividend
    events via /api/dividend-events/import. Rows with Type='MB' land as
    kind='managing_bonus'; Type='Dividend'/blank land as kind='payment'.
  - Any sheet whose name contains 'Share' or looks like a month
    (e.g. '2024 Dec', '2024 NOV') → per-shareholder share-allocation
    snapshots via /api/share-allocations.
  - Writes alias maps (company + shareholder) via /api/alias-maps so
    later imports don't need the same hand-mapping.

Usage
  # Dry-run (no writes)
  python3 scripts/import_dividend_xlsx.py \\
    --xlsx ~/Desktop/Dividend\\ 2025\\ 202601.xlsx \\
    --base-url http://localhost:8811 \\
    --email you@example.com --password ****

  # Commit
  python3 scripts/import_dividend_xlsx.py ... --commit
"""

from __future__ import annotations
import argparse
import datetime as _dt
import re
import sys
from typing import Optional

import openpyxl
import requests


COMPANY_ALIASES = {
    # Spreadsheet label → QBO company name OR legal_name.
    # Matcher checks both (case-insensitive).
    "SHG":         "Sweet Hut Group LLC",
    "Duluth":      "Sweet Hut Pleasant Hill",
    "Central":     "Sweet Hut Kitchen",
    "Ewe Group":   "Sweet Hut Doraville",
    "Texas":       "Sweet Hut Texas",
    "FT2":         "Food Terminal (West Midtown)",
    "Farm Noodle": "FARM NOODLE INC.",   # matches on legal_name (display name is "Food Terminal")
}

# Shareholder name aliases — left is what the spreadsheet says, right is
# the canonical display_name this tool uses when creating / looking up
# shareholder records.
SHAREHOLDER_ALIASES = {
    "Howie Ewe": "Howie",
    "Tao and Rachel": "Rachel and Tao",
    # Add more if the spreadsheet drifts from QBO / canonical records
}


# Month-header tabs to treat as allocation snapshots (effective at the
# end of the named month). '2024 Share' and 'Share Mar 23' / 'Share Jan
# 23' are also treated as snapshots.
MONTH_ABBR = {m: i for i, m in enumerate(
    ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"], start=1
)}


def _sheet_effective_date(name: str) -> Optional[str]:
    """Infer an effective_date (YYYY-MM-DD) from a share snapshot tab name.
    Returns None if the tab isn't a snapshot we understand."""
    n = name.strip()
    # "2024 Dec", "2024 NOV"
    m = re.match(r"^(\d{4})\s+([A-Za-z]{3,})$", n)
    if m:
        year = int(m.group(1))
        abbr = m.group(2)[:3].upper()
        month = MONTH_ABBR.get(abbr)
        if month:
            # Last day of the month
            if month == 12:
                return f"{year}-12-31"
            nxt = _dt.date(year, month + 1, 1) - _dt.timedelta(days=1)
            return nxt.strftime("%Y-%m-%d")
    # "2024 AUG-OCT" → take the last month of the range
    m = re.match(r"^(\d{4})\s+([A-Za-z]{3,})-([A-Za-z]{3,})$", n)
    if m:
        year = int(m.group(1))
        month = MONTH_ABBR.get(m.group(3)[:3].upper())
        if month:
            if month == 12:
                return f"{year}-12-31"
            nxt = _dt.date(year, month + 1, 1) - _dt.timedelta(days=1)
            return nxt.strftime("%Y-%m-%d")
    # "2024 Share" → Dec 31 of that year
    m = re.match(r"^(\d{4})\s+Share$", n)
    if m:
        return f"{m.group(1)}-12-31"
    # "Share Mar 23", "Share Jan 23" → end of that month, year = 2000+DD
    m = re.match(r"^Share\s+([A-Za-z]{3,})\s+(\d{2})$", n)
    if m:
        month = MONTH_ABBR.get(m.group(1)[:3].upper())
        yr = 2000 + int(m.group(2))
        if month:
            if month == 12:
                return f"{yr}-12-31"
            nxt = _dt.date(yr, month + 1, 1) - _dt.timedelta(days=1)
            return nxt.strftime("%Y-%m-%d")
    return None


class Client:
    """Minimal wrapper around the Consolidated Report API."""
    def __init__(self, base_url: str, token: str):
        self.base = base_url.rstrip("/")
        self.h = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    @classmethod
    def login(cls, base_url: str, email: str, password: str) -> "Client":
        r = requests.post(
            base_url.rstrip("/") + "/api/auth/login",
            json={"email": email, "password": password},
            timeout=15,
        )
        r.raise_for_status()
        return cls(base_url, r.json()["token"])

    def _get(self, path: str, **params):
        r = requests.get(self.base + path, headers=self.h, params=params, timeout=30)
        r.raise_for_status()
        return r.json()

    def _post(self, path: str, body: dict):
        r = requests.post(self.base + path, headers=self.h, json=body, timeout=60)
        if not r.ok:
            raise RuntimeError(f"POST {path} → {r.status_code} {r.text[:400]}")
        return r.json() if r.text else None

    def companies(self):
        return self._get("/api/companies")

    def shareholders(self):
        return self._get("/api/shareholders", include_inactive=True)

    def create_shareholder(self, display_name: str, short_name: Optional[str] = None):
        return self._post("/api/shareholders", {"display_name": display_name, "short_name": short_name, "active": True})

    def upsert_alias(self, kind: str, source_label: str, target_id: str):
        return self._post("/api/alias-maps", {"alias_kind": kind, "source_label": source_label, "target_id": target_id})

    def upsert_allocation(self, shareholder_id: str, effective_date: str, shares_held: float,
                          ownership_pct: Optional[float], dps: Optional[float], mb: Optional[float],
                          notes: Optional[str] = None):
        return self._post("/api/share-allocations", {
            "shareholder_id": shareholder_id,
            "effective_date": effective_date,
            "shares_held": shares_held,
            "ownership_pct": ownership_pct,
            "dividend_per_share": dps,
            "mb_amount": mb or 0,
            "notes": notes,
        })

    def import_events(self, rows: list, commit: bool):
        return self._post("/api/dividend-events/import", {"rows": rows, "commit": commit})

    def create_event(self, body: dict):
        return self._post("/api/dividend-events", body)


def _match_company(companies: list, label: str):
    wanted = (COMPANY_ALIASES.get(label) or label).strip().lower()
    if not wanted:
        return None
    for c in companies:
        if (c.get("name") or "").strip().lower() == wanted:
            return c
        if (c.get("legal_name") or "").strip().lower() == wanted:
            return c
    return None


def _match_or_create_shareholder(client: Client, cache: dict, source_name: str, dry_run: bool):
    canonical = SHAREHOLDER_ALIASES.get(source_name, source_name).strip()
    key = canonical.lower()
    if key in cache:
        return cache[key]
    # Miss — create unless dry-run
    if dry_run:
        placeholder = {"id": None, "display_name": canonical, "_planned": True}
        cache[key] = placeholder
        return placeholder
    created = client.create_shareholder(canonical, short_name=canonical.split()[0])
    cache[key] = created
    return created


def _parse_data_sheet(ws, companies: list, sh_cache: dict, client: Client, dry_run: bool):
    """Turn the `Data` sheet into importable event rows."""
    headers = [c.value for c in ws[1]]
    H = {h: i for i, h in enumerate(headers) if h}
    required = ["Shareholder", "Type", "Dividend Determined", "Dividend Issued", "Company Source", "Date"]
    missing = [k for k in required if k not in H]
    if missing:
        raise SystemExit(f"Data sheet missing columns: {missing}")

    rows_to_import = []
    issues = []
    unknown_companies = set()
    skipped_no_date = 0
    skipped_no_amount = 0

    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r[0] is None and r[1] is None:  # blank row
            continue
        sh_label = r[H["Shareholder"]]
        if not sh_label:
            continue
        co_label = r[H["Company Source"]]
        type_label = r[H["Type"]] or "Dividend"
        iss = r[H["Dividend Issued"]]
        dt = r[H["Date"]]
        method = r[H.get("Method", -1)] if "Method" in H else None
        period = r[H.get("Period", -1)] if "Period" in H else None

        if iss in (None, 0):
            skipped_no_amount += 1
            continue
        try:
            amount = float(iss)
        except (TypeError, ValueError):
            issues.append({"row": i, "error": "non-numeric amount", "raw": iss})
            continue
        if amount <= 0:
            continue

        if not dt:
            skipped_no_date += 1
            issues.append({"row": i, "error": "missing date", "sh": sh_label, "amt": amount})
            continue
        if isinstance(dt, _dt.datetime):
            dt_str = dt.strftime("%Y-%m-%d")
        elif isinstance(dt, _dt.date):
            dt_str = dt.strftime("%Y-%m-%d")
        elif isinstance(dt, str):
            dt_str = dt[:10]
        else:
            issues.append({"row": i, "error": "bad date type", "raw": repr(dt)})
            continue

        co = _match_company(companies, str(co_label or ""))
        if not co:
            unknown_companies.add(co_label)
            issues.append({"row": i, "error": "unknown company", "label": co_label})
            continue

        sh = _match_or_create_shareholder(client, sh_cache, str(sh_label), dry_run)

        kind = "managing_bonus" if (str(type_label).strip().upper() == "MB") else "payment"
        memo_bits = [p for p in [period, method] if p]
        memo = " · ".join(str(x) for x in memo_bits) or None

        rows_to_import.append({
            "event_date": dt_str,
            "shareholder_id_hint": sh.get("id"),
            "shareholder_source": sh_label,
            "company_id": co["id"],
            "company_name": co["name"],
            "amount": round(amount, 2),
            "kind": kind,
            "memo": memo,
        })

    return {
        "rows": rows_to_import,
        "issues": issues,
        "unknown_companies": sorted(unknown_companies),
        "skipped_no_date": skipped_no_date,
        "skipped_no_amount": skipped_no_amount,
    }


def _parse_share_sheet(ws, eff_date: str):
    """Share-allocation sheets have a fixed structure (column B onward)."""
    out = []
    # Col A: name; B: ownership fraction; C: shares; D: dividend; E: MB; F: total; G: effective %
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None: continue
        name = str(row[0]).strip()
        # Skip summary/total rows
        if name.startswith(("Total", "Grand")) or any(
            name.lower().startswith(p) for p in ("total ", "dps", "total shares")
        ):
            continue
        # Columns
        own = row[1] if len(row) > 1 else None
        shares = row[2] if len(row) > 2 else None
        div = row[3] if len(row) > 3 else None
        mb = row[4] if len(row) > 4 else None
        if shares in (None, 0) and not div:
            continue
        try:
            s = float(shares or 0)
            o = float(own) if own is not None else None
            d_amt = float(div or 0)
            mb_amt = float(mb or 0)
        except (TypeError, ValueError):
            continue
        dps = (d_amt / s) if s > 0 and d_amt > 0 else None
        out.append({
            "name": name,
            "effective_date": eff_date,
            "shares_held": s,
            "ownership_pct": o,
            "dividend_per_share": dps,
            "mb_amount": mb_amt,
        })
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--base-url", default="http://localhost:8811")
    ap.add_argument("--email", required=True)
    ap.add_argument("--password", required=True)
    ap.add_argument("--commit", action="store_true", help="Actually write. Without this, dry-run only.")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    client = Client.login(args.base_url, args.email, args.password)

    # Prefetch
    companies = client.companies()
    existing = client.shareholders()
    sh_cache = {(s["display_name"] or "").lower(): s for s in existing}
    # Also seed from aliases so a spreadsheet label finds an existing canonical record
    for src, canonical in SHAREHOLDER_ALIASES.items():
        if canonical.lower() in sh_cache:
            sh_cache[src.lower()] = sh_cache[canonical.lower()]

    # Alias maps (company side) — upsert so the audit trail exists
    company_aliases_to_write = []
    for label, qbo_name in COMPANY_ALIASES.items():
        co = next((c for c in companies if (c.get("name") or "").strip().lower() == qbo_name.lower()), None)
        if co:
            company_aliases_to_write.append(("company", label, co["id"]))

    # ---- Phase 1: share allocation snapshots ----
    share_sheets = [(n, _sheet_effective_date(n)) for n in wb.sheetnames]
    share_sheets = [(n, d) for (n, d) in share_sheets if d]
    # Oldest first so DPS/MB history is built up chronologically
    share_sheets.sort(key=lambda x: x[1])

    allocation_rows = []
    for (name, eff) in share_sheets:
        allocation_rows.append((name, eff, _parse_share_sheet(wb[name], eff)))

    # ---- Phase 2: data sheet events ----
    parsed = _parse_data_sheet(wb["Data"], companies, sh_cache, client, dry_run=not args.commit)

    # ---- Print summary ----
    print(f"\n═══ DRY-RUN SUMMARY (--commit not set)" if not args.commit else "\n═══ COMMIT SUMMARY")
    print(f"\nConnected QBO companies:          {len(companies)}")
    print(f"Existing shareholders:            {len(existing)}")
    print(f"Company aliases to write:         {len(company_aliases_to_write)}")
    print(f"\nShare allocation sheets found:    {len(share_sheets)}")
    for (name, eff, rows) in allocation_rows:
        print(f"  - {name:20s} ({eff})  {len(rows)} shareholders")

    print(f"\nData sheet events to import:      {len(parsed['rows'])}")
    print(f"  ↳ payment (pro-rata Dividend):  {sum(1 for r in parsed['rows'] if r['kind']=='payment')}")
    print(f"  ↳ managing_bonus (MB):          {sum(1 for r in parsed['rows'] if r['kind']=='managing_bonus')}")
    print(f"  Skipped rows missing date:      {parsed['skipped_no_date']}")
    print(f"  Skipped rows with no amount:    {parsed['skipped_no_amount']}")
    print(f"  Other issues:                   {len(parsed['issues'])}")
    if parsed["unknown_companies"]:
        print(f"  UNKNOWN company labels:         {parsed['unknown_companies']}")

    # Show first few issues
    for iss in parsed["issues"][:10]:
        print(f"    ! {iss}")

    # Sum by shareholder (sanity check)
    by_sh = {}
    for r in parsed["rows"]:
        by_sh.setdefault(r["shareholder_source"], 0)
        by_sh[r["shareholder_source"]] += r["amount"]
    if by_sh:
        print("\nTotals by shareholder (importable rows):")
        for n, t in sorted(by_sh.items(), key=lambda x: -x[1]):
            print(f"  ${t:>10,.2f}   {n}")

    total_importable = sum(r["amount"] for r in parsed["rows"])
    print(f"\nTotal $ in importable events:     ${total_importable:,.2f}")

    if not args.commit:
        print("\nNo writes made. Re-run with --commit to apply.")
        return

    # ---- Commit ----
    # Phase A: upsert shareholders
    for r in parsed["rows"]:
        sh = _match_or_create_shareholder(client, sh_cache, r["shareholder_source"], dry_run=False)
        r["shareholder_id_hint"] = sh["id"]
    # Allocations too (in case they reference shareholders not in Data)
    for (_, _, rows) in allocation_rows:
        for r in rows:
            _match_or_create_shareholder(client, sh_cache, r["name"], dry_run=False)

    # Phase B: aliases (shareholder + company)
    written_aliases = 0
    for kind, label, tid in company_aliases_to_write:
        try:
            client.upsert_alias(kind, label, tid)
            written_aliases += 1
        except Exception as e:
            print(f"  alias(company,{label}) failed:", e)
    for src, canonical in SHAREHOLDER_ALIASES.items():
        sh = sh_cache.get(canonical.lower())
        if sh and sh.get("id"):
            try:
                client.upsert_alias("shareholder", src, sh["id"])
                written_aliases += 1
            except Exception as e:
                print(f"  alias(shareholder,{src}) failed:", e)
    print(f"\nAliases written:                   {written_aliases}")

    # Phase C: share allocations
    written_alloc = 0
    for (name, eff, rows) in allocation_rows:
        for r in rows:
            sh = sh_cache.get(SHAREHOLDER_ALIASES.get(r["name"], r["name"]).lower())
            if not sh or not sh.get("id"): continue
            try:
                client.upsert_allocation(
                    sh["id"], eff,
                    shares_held=r["shares_held"],
                    ownership_pct=r["ownership_pct"],
                    dps=r["dividend_per_share"],
                    mb=r["mb_amount"],
                    notes=name,
                )
                written_alloc += 1
            except Exception as e:
                print(f"  allocation({name},{r['name']}) failed:", e)
    print(f"Allocations written:               {written_alloc}")

    # Phase D: events — use single-event API so we can specify `kind` +
    # store memo + attribute posting status as 'skipped' (already in QBO).
    written_evt = 0
    for r in parsed["rows"]:
        body = {
            "shareholder_id": r["shareholder_id_hint"],
            "company_id": r["company_id"],
            "event_date": r["event_date"],
            "amount": r["amount"],
            "kind": r["kind"],
            "memo": r["memo"],
            "post_to_qbo": False,  # historical import — JE already exists in QBO
        }
        try:
            client.create_event(body)
            written_evt += 1
        except Exception as e:
            print(f"  event({r['event_date']},{r['shareholder_source']},${r['amount']}) failed:", e)
    print(f"Events written:                    {written_evt}")
    print("\nDone.")


if __name__ == "__main__":
    main()
